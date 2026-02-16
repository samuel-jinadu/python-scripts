import sys, openpyxl, os
from openpyxl.utils import get_column_letter

# changing directory

os.chdir(r"C:\Users\Samuel\Desktop\Remote Job\practice\python\multiplicationTable")

# getting the N

N = 6

if len(sys.argv) < 2:
	print("The default is '6'", "\n")
else:
	N = int(sys.argv[1])


# excel stuff
print("\n" + "Making workbook...", "\n")
wb = openpyxl.Workbook()
sheet = wb.active

for row in range(1, N + 1):
	for column in range(1, N + 1):
		sheet[get_column_letter(column) + str(row)] = row * column

sheet.freeze_panes = "B2"

print("Saving workbook...", "\n")
filename = f"multiplicationTableFor{N}.xlsx"
wb.save(filename)

print(f"Workbook saved at {os.getcwd()}\\{filename}", "\n")
